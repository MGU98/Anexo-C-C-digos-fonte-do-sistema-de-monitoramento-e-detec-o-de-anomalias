
import json
import argparse
from pathlib import Path
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import load_workbook, Workbook

class SensorDataHandler(BaseHTTPRequestHandler):

    def __init__(self, output_dir, *args, **kwargs):
        self.output_dir = output_dir
        super().__init__(*args, **kwargs)

    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"1")

    def do_POST(self):
        try:
            content_length = int(self.headers["Content-Length"])
            post_data = self.rfile.read(content_length).decode("utf-8")
            sensor_data = json.loads(post_data)
            print('DATA BEGIN')
            print(post_data)
            print(sensor_data)
            print('DATA END')

            filepath = Path(self.output_dir) / "sensor_data_values.xlsx"
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            self._save_data_to_xlsx(sensor_data, filepath, timestamp)

            print(f"Data appended to {filepath}")
            self.send_response(204)

        except Exception as e:
            print(f"Error processing data: {str(e)}")
            self.send_response(500)

        self.end_headers()

    def _save_data_to_xlsx(self, data, filepath: Path, timestamp: str):

      # Check if file exists and load or create workbook
      if filepath.exists():
          workbook = load_workbook(filename=filepath)
          sheet = workbook.active
          print("Workbook loaded.")
      else:
          print(f"File does not exist: {filepath}, creating new workbook.")
          workbook = Workbook()
          sheet = workbook.active
          sheet.append(["timestamp", "x", "y", "z", "current", "has_anomaly"])  # Write header

      # Append data rows
      samples = len(data["x"])
      for i in range(samples):
          sheet.append([
              timestamp,
              data["x"][i],
              data["y"][i],
              data["z"][i],
              data["current"][i],
              False
          ])

      # Save workbook
      workbook.save(filepath)
      print(f"Data written to: {filepath}")

def create_server(output_dir, port):

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    def handler(*args, **kwargs):
        return SensorDataHandler(output_dir, *args, **kwargs)

    return HTTPServer(("", port), handler)

def main():
    parser = argparse.ArgumentParser(description="Sensor Data Collection Server")
    parser.add_argument(
        "-d",
        "--dir",
        type=str,
        default="sensor_data",
        help="Output directory for sensor data (default: sensor_data)",
    )
    parser.add_argument(
        "-p", "--port", type=int, default=4242, help="Server port (default: 4242)"
    )
    args = parser.parse_args()

    server = create_server(args.dir, args.port)

    print("\nSensor Data Collection Server")
    print(f"Saving data to: {args.dir}")
    print(f"Server running on port {args.port}")
    print("Press Ctrl+C to stop\n")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer shutting down...")
        server.server_close()

if __name__ == "__main__":
    main()
